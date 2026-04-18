import os
import json
import re
import requests
import tempfile
from fastapi import FastAPI, Request
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
OPENROUTER_KEY = os.getenv("OPENROUTER_KEY")
ADMIN_CHAT_ID = os.getenv("ADMIN_CHAT_ID")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None
app = FastAPI()

conversations = {}
user_settings = {}
DEFAULT_BUSINESS_ID = "1651e4c3-0215-4f04-abd3-68c7dba3e380"

def get_business_id(context_business_id=None):
    business_id = context_business_id if context_business_id else DEFAULT_BUSINESS_ID
    print("Using business_id:", business_id)
    return business_id

def fmt(val):
    try:
        if val is None: return 0
        f = float(val)
        return int(f) if f.is_integer() else f
    except:
        return val

def send_message(chat_id, text):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": chat_id, "text": text}
    requests.post(url, json=payload)

def ask_ai(text: str, pending_context: dict = None, photo_url: str = None, strict_mode: bool = False) -> str:
    if strict_mode:
        system_prompt = 'Extract transactions ONLY. Return STRICT JSON. No explanation. Format: {"transactions": [{"amount": number, "type": "income|expense|liability", "category": "string", "note": "original text"}]}'
    else:
        system_prompt = """You are a Chartered Accountant (CA) and financial advisor for an ecommerce company called De Markt.

Your responsibilities:
- Understand natural language financial inputs (English, Bangla, mixed)
- Extract transactions correctly
- Classify into: income, expense, liability
- Categorize properly (sales, rent, salary, transport, vendor_due, etc.)
- Manage vendor dues, salaries, ads, and expenses
- Track cash, bank, and liabilities
- Ask follow-up questions when needed
- Behave like a human finance manager (not a robot)

IMPORTANT RULES:
1. Always be short, clear, professional, and human.
2. Understand:
- sales = income
- rent / salary / ads = expense
- borrowed / supplier due = liability
3. If the user's setup or onboarding reply is ambiguous (e.g. combining bank and cash as one number), explicitly ask: "Please break it down: cash, bank, and bkash separately."
4. If multiple lines are sent, extract multiple transactions."""

    context_str = f"\n\nPending Context: {json.dumps(pending_context)}" if pending_context else ""
    user_content = []
    if text: user_content.append({"type": "text", "text": text})
    if photo_url: user_content.append({"type": "image_url", "image_url": {"url": photo_url}})
    
    if not user_content: return "{}"

    try:
        res = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENROUTER_KEY}"},
            json={
                "model": "openai/gpt-4o-mini",
                "messages": [
                    {"role": "system", "content": system_prompt + context_str},
                    {"role": "user", "content": user_content}
                ]
            }
        )
        res.raise_for_status()
        raw_content = res.json()["choices"][0]["message"]["content"].strip()
        if raw_content.startswith("```json"): raw_content = raw_content[7:]
        if raw_content.endswith("```"): raw_content = raw_content[:-3]
        return raw_content.strip()
    except Exception as e:
        return "{}"

@app.get("/cron")
async def run_cron():
    targets = list(user_settings.keys())
    if ADMIN_CHAT_ID and int(ADMIN_CHAT_ID) not in targets:
        targets.append(int(ADMIN_CHAT_ID))
    for cid in targets:
        send_message(cid, "De Markt day-end update\n\nReply shortly:\n- total sales?\n- any vendor purchase/payment?\n- total expenses?\n- any salary paid?\n- ad spend today?\n- founder withdrawal/investment?\n- any bills paid?")
    return {"ok": True}

@app.post("/webhook")
async def webhook(request: Request):
    try: update = await request.json()
    except: return {"ok": True}

    msg = update.get("message", {})
    chat_id = msg.get("chat", {}).get("id")
    text = str(msg.get("text", msg.get("caption", ""))).strip()

    if not chat_id: return {"ok": True}

    try:
        print(f"DEBUG INCOMING: Chat: {chat_id} | Text: {text}")
        
        # Load Memory Profile implicitly
        pending = conversations.get(chat_id, {"history": [], "onboarding_mode": False})
        
        trigger_words = ["you are hired", "start", "setup", "manage my finance"]
        if any(tw in text.lower() for tw in trigger_words):
            if not pending.get("onboarding_mode"):
                pending["onboarding_mode"] = True
                print("DEBUG: Activating implicit onboarding mode")

        # Document check for Excel processing
        document = msg.get("document", {})
        if document.get("file_name", "").endswith(".xlsx"):
            import openpyxl
            f_res = requests.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={document['file_id']}").json()
            if f_res.get("ok"):
                send_message(chat_id, "Processing Excel file. Standby...")
                doc_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{f_res['result']['file_path']}"
                r = requests.get(doc_url)
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(r.content)
                    tmp_path = tmp.name
                try:
                    wb = openpyxl.load_workbook(tmp_path, data_only=True)
                    ws = wb.active
                    headers = [str(c.value).strip().lower() for c in ws[1] if c.value]
                    products_saved = 0
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0]: continue
                        p_data = dict(zip(headers, row))
                        name = str(p_data.get("product_name", row[0]))
                        if supabase:
                            supabase.table("transactions").insert({
                                "business_id": get_business_id(),
                                "type": "product_master",
                                "category": name,
                                "amount": float(p_data.get("selling_price", 0) or 0),
                                "note": json.dumps({
                                    "sku": str(p_data.get("sku", "")),
                                    "cost_price": float(p_data.get("cost_price", 0) or 0),
                                    "selling_price": float(p_data.get("selling_price", 0) or 0),
                                    "stock_qty": int(p_data.get("stock_qty", 0) or 0),
                                    "status": str(p_data.get("status", "active"))
                                }),
                                "source": "telegram"
                            }).execute()
                            products_saved += 1
                    send_message(chat_id, f"Successfully imported {products_saved} products to memory.")
                except Exception as e:
                    send_message(chat_id, f"Excel Import Error: {str(e)}")
                os.remove(tmp_path)
                return {"ok": True}

        photo_url = None
        if "photo" in msg:
            file_id = msg["photo"][-1]["file_id"]
            f_res = requests.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={file_id}").json()
            if f_res.get("ok"):
                photo_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{f_res['result']['file_path']}"

        if chat_id not in user_settings: user_settings[chat_id] = "22:00"

        # Command Execution
        if text.startswith("/"):
            commands = [c.strip().lower() for c in re.split(r'[,\n]+', text) if c.strip()]
            responses = []
            try:
                for cmd in commands:
                    if cmd == "/products":
                        if supabase:
                            res = supabase.table("transactions").select("category, amount, note").eq("type", "product_master").order("created_at", desc=False).execute()
                            active_items = {}
                            for r in res.data: active_items[r['category']] = r
                            s = "Product Database:\n"
                            for k, v in active_items.items(): s += f"- {k} (Selling: {fmt(v['amount'])})\n"
                            responses.append(s)
                    elif cmd.startswith("/product"):
                        parts = cmd.split(maxsplit=1)
                        if len(parts) > 1 and supabase:
                            res = supabase.table("transactions").select("*").eq("type", "product_master").eq("category", parts[1]).order("created_at", desc=True).limit(1).execute()
                            if res.data:
                                n = json.loads(res.data[0].get("note", "{}"))
                                responses.append(f"Product: {res.data[0]['category']}\nSelling Price: {fmt(res.data[0]['amount'])}\nCost Price: {fmt(n.get('cost_price',0))}\nStock: {n.get('stock_qty',0)}\nSKU: {n.get('sku','')}")
                            else: responses.append("Product not found.")
                    elif cmd == "/summary":
                        if supabase:
                            transactions = supabase.table("transactions").select("*").execute().data
                            income = sum(float(t.get("amount") or 0) for t in transactions if t.get("type") == "income")
                            expense = sum(float(t.get("amount") or 0) for t in transactions if t.get("type") == "expense")
                            liability = sum(float(t.get("amount") or 0) for t in transactions if t.get("type") == "liability")
                            responses.append(f"Summary\nIncome: {fmt(income)}\nExpense: {fmt(expense)}\nLiability: {fmt(liability)}\nBalance: {fmt(income-expense)}")
            except Exception as e:
                 responses.append(f"Command Error: {str(e)}")
            if responses: send_message(chat_id, "\n\n".join(responses))
            return {"ok": True}

        # STEP 1 — RECEIVE MESSAGE
        user_text = text.lower().strip()
        if not user_text:
            send_message(chat_id, "Please send something.")
            return {"ok": True}

        # STEP 2 — SIMPLE DETECTION (NO AI)
        has_number = any(char.isdigit() for char in user_text)

        # STEP 3 — TRANSACTION MODE (PRIORITY)
        if has_number:
            lines = [line.strip() for line in user_text.split('\n') if line.strip()]
            reply_lines = []
            
            for line in lines:
                if not any(char.isdigit() for char in line):
                    continue
                    
                num_match = re.search(r'\d+(\.\d+)?', line)
                amount = float(num_match.group()) if num_match else 0.0
                
                tx_type, category = "expense", "general"
                if "sales" in line or "sale" in line:
                    tx_type, category = "income", "sales"
                elif "rent" in line:
                    tx_type, category = "expense", "rent"
                elif "salary" in line:
                    tx_type, category = "expense", "salary"
                elif "transport" in line:
                    tx_type, category = "expense", "transport"
                elif "bought" in line:
                    tx_type, category = "expense", "equipment"
                elif "borrowed" in line:
                    tx_type, category = "liability", "loan"
                elif "due" in line:
                    tx_type, category = "liability", "supplier_due"
                
                tx_type = tx_type.lower()
                
                is_valid = (
                    isinstance(amount, (int, float)) 
                    and tx_type in ["income", "expense", "liability"] 
                    and isinstance(category, str) and len(category) > 0
                )

                if is_valid:
                    if supabase:
                        print("USER TEXT:", user_text)
                        print("AMOUNT:", amount)
                        print("TYPE:", tx_type)
                        print("CATEGORY:", category)
                        
                        payload = {
                            "business_id": "1651e4c3-0215-4f04-abd3-68c7dba3e380",
                            "amount": float(amount),
                            "type": tx_type.lower(),
                            "category": category,
                            "note": user_text,
                            "source": "telegram"
                        }
                        
                        print("PAYLOAD:", payload)
                        try:
                            response = supabase.table("transactions").insert(payload).execute()
                            print("SUPABASE SUCCESS:", response)
                            reply = f"Saved\nAmount: {amount}\nType: {tx_type}\nCategory: {category}"
                            reply_lines.append(reply)
                        except Exception as e:
                            print("SUPABASE INSERT ERROR:", repr(e))
                            reply = f"Saving failed. Error logged."
                            reply_lines.append(reply)
                else:
                    reply_lines.append("Saving failed. Check logs.")
            
            if reply_lines:
                send_message(chat_id, "\n\n".join(reply_lines))
            return {"ok": True}

        # STEP 4 — BASIC COMMANDS
        if user_text in ["hi", "hello"]:
            send_message(chat_id, "I'm managing your finances. Tell me a transaction or ask a question.")
            return {"ok": True}
            
        if "hired" in user_text:
            send_message(chat_id, "Got it. I'll act as your CA. Let's start with your current cash/bank balance.")
            return {"ok": True}

        # STEP 5 — SIMPLE QUESTIONS (NO AI YET)
        if "cash" in user_text or "balance" in user_text:
            balance_val = 0
            if supabase:
                transactions = supabase.table("transactions").select("amount, type").execute().data
                income = sum(float(t.get("amount") or 0) for t in transactions if t.get("type") == "income")
                expense = sum(float(t.get("amount") or 0) for t in transactions if t.get("type") == "expense")
                balance_val = income - expense
            send_message(chat_id, f"Current balance is {balance_val}")
            return {"ok": True}

        if "vendor" in user_text and "due" in user_text:
            due_val = 0
            if supabase:
                txs = supabase.table("transactions").select("amount, type, category").execute().data
                due_val = sum(float(t.get("amount") or 0) for t in txs if t.get("type") == "liability" and t.get("category") in ["supplier_due", "due"])
            send_message(chat_id, f"Total vendor due is {due_val}")
            return {"ok": True}

        # STEP 6 — FALLBACK (LAST ONLY)
        send_message(chat_id, "I didn’t understand. Try like: sales 5000 or rent 1200")
        return {"ok": True}

    except Exception as master_err:
        print(f"DEBUG CRITICAL APP FAILURE: {str(master_err)}")
        send_message(chat_id, "System error. Please send again.")
        return {"ok": True}